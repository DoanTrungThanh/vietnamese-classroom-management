import io
from flask import make_response
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import urllib.parse

def create_excel_response(data, filename, sheet_name='Sheet1'):
    """
    Create Excel file response from data with UTF-8 encoding and .xlsx format
    """
    try:
        # Create Excel file in memory
        output = io.BytesIO()

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add headers if data exists
        if data:
            headers = list(data[0].keys())
            ws.append(headers)

            # Add data rows
            for row_data in data:
                row = [row_data.get(header, '') for header in headers]
                ws.append(row)

        # Style the header row if data exists
        if data and ws.max_row > 0:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Apply header styling to first row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Data rows
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Save to BytesIO with UTF-8 encoding
        wb.save(output)
        output.seek(0)

        # Ensure filename is UTF-8 encoded and properly formatted
        if not filename.endswith('.xlsx'):
            filename = filename.replace('.xls', '.xlsx')
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

        # URL encode filename for proper UTF-8 handling
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))

        # Create response with proper headers for UTF-8 and .xlsx
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        response.headers['Cache-Control'] = 'no-cache'

        return response

    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        return None

def export_users_to_excel(users):
    """
    Export users data to Excel
    """
    data = []
    for user in users:
        data.append({
            'ID': user.id,
            'Họ và tên': user.full_name,
            'Tên đăng nhập': user.username,
            'Email': user.email,
            'Số điện thoại': user.phone or '',
            'Vai trò': {
                'admin': 'Quản trị viên',
                'manager': 'Quản sinh',
                'teacher': 'Giáo viên'
            }.get(user.role, user.role),
            'Trạng thái': 'Hoạt động' if user.is_active else 'Không hoạt động',
            'Ngày tạo': user.created_at.strftime('%d/%m/%Y %H:%M') if user.created_at else '',
            'Địa chỉ': user.address or ''
        })
    
    filename = f'danh_sach_nguoi_dung_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Danh sách người dùng')

def export_classes_to_excel(classes):
    """
    Export classes data to Excel
    """
    data = []
    for class_obj in classes:
        data.append({
            'ID': class_obj.id,
            'Tên lớp': class_obj.name,
            'Mô tả': class_obj.description or '',
            'Quản sinh': class_obj.manager.full_name if class_obj.manager else '',
            'Sĩ số hiện tại': class_obj.student_count,
            'Sĩ số tối đa': class_obj.max_students or '',
            'Trạng thái': 'Hoạt động' if class_obj.is_active else 'Không hoạt động',
            'Ngày tạo': class_obj.created_at.strftime('%d/%m/%Y %H:%M') if class_obj.created_at else ''
        })
    
    filename = f'danh_sach_lop_hoc_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Danh sách lớp học')

def export_students_to_excel(students):
    """
    Export students data to Excel
    """
    data = []
    for student in students:
        data.append({
            'ID': student.id,
            'Họ và tên': student.full_name,
            'Mã học sinh': student.student_id or '',
            'Lớp học': student.class_obj.name if student.class_obj else '',
            'Ngày sinh': student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '',
            'Tên phụ huynh': student.parent_name or '',
            'SĐT phụ huynh': student.parent_phone or '',
            'Địa chỉ': student.address or '',
            'Trạng thái': 'Đang học' if student.is_active else 'Đã nghỉ',
            'Ngày nhập học': student.enrollment_date.strftime('%d/%m/%Y') if student.enrollment_date else ''
        })
    
    filename = f'danh_sach_hoc_sinh_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Danh sách học sinh')

def export_expenses_to_excel(expenses):
    """
    Export expenses data to Excel
    """
    data = []
    for expense in expenses:
        data.append({
            'ID': expense.id,
            'Tiêu đề': expense.title,
            'Mô tả': expense.description or '',
            'Số tiền': float(expense.amount),
            'Ngày chi tiêu': expense.expense_date.strftime('%d/%m/%Y'),
            'Danh mục': expense.category.name,
            'Nhà cung cấp': expense.vendor or '',
            'Số hóa đơn': expense.receipt_number or '',
            'Phương thức thanh toán': expense.payment_method_display,
            'Trạng thái': expense.status_display,
            'Người tạo': expense.creator.full_name if expense.creator else '',
            'Người duyệt': expense.approver.full_name if expense.approver else '',
            'Ngày tạo': expense.created_at.strftime('%d/%m/%Y %H:%M') if expense.created_at else '',
            'Ghi chú': expense.notes or ''
        })
    
    filename = f'danh_sach_chi_tieu_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Danh sách chi tiêu')

def export_attendance_to_excel(attendance_records, class_name=None, date_range=None):
    """
    Export attendance data to Excel
    """
    data = []
    for record in attendance_records:
        data.append({
            'ID': record.id,
            'Học sinh': record.student.full_name,
            'Lớp': record.schedule.class_obj.name,
            'Giáo viên': record.schedule.teacher.full_name,
            'Ngày': record.date.strftime('%d/%m/%Y'),
            'Khung giờ': record.schedule.time_slot.name,
            'Thời gian': f"{record.schedule.time_slot.start_time.strftime('%H:%M')} - {record.schedule.time_slot.end_time.strftime('%H:%M')}",
            'Trạng thái': {
                'present': 'Có mặt',
                'absent_without_reason': 'Vắng không phép',
                'absent_with_reason': 'Vắng có phép'
            }.get(record.status, record.status),
            'Ghi chú': record.notes or ''
        })
    
    # Create filename with class and date info
    filename_parts = ['diem_danh']
    if class_name:
        filename_parts.append(class_name.replace(' ', '_'))
    if date_range:
        filename_parts.append(date_range.replace('/', ''))
    filename_parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    
    filename = f'{"_".join(filename_parts)}.xlsx'
    return create_excel_response(data, filename, 'Điểm danh')

def export_schedule_to_excel(schedules):
    """
    Export schedule data to Excel
    """
    data = []
    for schedule in schedules:
        data.append({
            'ID': schedule.id,
            'Giáo viên': schedule.teacher.full_name,
            'Lớp học': schedule.class_obj.name,
            'Khung giờ': schedule.time_slot.name,
            'Thời gian': f"{schedule.time_slot.start_time.strftime('%H:%M')} - {schedule.time_slot.end_time.strftime('%H:%M')}",
            'Thứ': {
                0: 'Chủ nhật',
                1: 'Thứ 2',
                2: 'Thứ 3', 
                3: 'Thứ 4',
                4: 'Thứ 5',
                5: 'Thứ 6',
                6: 'Thứ 7'
            }.get(schedule.day_of_week, f'Thứ {schedule.day_of_week}'),
            'Trạng thái': 'Hoạt động' if schedule.is_active else 'Tạm dừng',
            'Ngày tạo': schedule.created_at.strftime('%d/%m/%Y %H:%M') if schedule.created_at else ''
        })
    
    filename = f'lich_day_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Lịch dạy')

def export_time_slots_to_excel(time_slots):
    """
    Export time slots data to Excel
    """
    data = []
    for slot in time_slots:
        data.append({
            'ID': slot.id,
            'Tên khung giờ': slot.name,
            'Thời gian bắt đầu': slot.start_time.strftime('%H:%M'),
            'Thời gian kết thúc': slot.end_time.strftime('%H:%M'),
            'Buổi': slot.session,
            'Mô tả': slot.description or '',
            'Trạng thái': 'Hoạt động' if slot.is_active else 'Không hoạt động',
            'Ngày tạo': slot.created_at.strftime('%d/%m/%Y %H:%M') if slot.created_at else ''
        })
    
    filename = f'khung_gio_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Khung giờ')

def export_financial_transactions_to_excel(transactions):
    """
    Export financial transactions data to Excel
    """
    data = []
    for transaction in transactions:
        data.append({
            'ID': transaction.id,
            'Loại giao dịch': 'Thu' if transaction.type == 'income' else 'Chi',
            'Số tiền (VNĐ)': f"{transaction.amount:,.0f}",
            'Mô tả': transaction.description or '',
            'Danh mục': transaction.category or '',
            'Lớp học': transaction.class_obj.name if transaction.class_obj else '',
            'Sự kiện': transaction.event.title if transaction.event else '',
            'Ngày giao dịch': transaction.transaction_date.strftime('%d/%m/%Y'),
            'Người tạo': transaction.creator.full_name if transaction.creator else '',
            'Ngày tạo': transaction.created_at.strftime('%d/%m/%Y %H:%M') if transaction.created_at else ''
        })

    filename = f'giao_dich_tai_chinh_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Giao dịch tài chính')

def export_donations_to_excel(donations):
    """
    Export donations data to Excel
    """
    data = []
    for donation in donations:
        data.append({
            'ID': donation.id,
            'Tên tài sản': donation.asset_name,
            'Người quyên góp': donation.donor_name or 'Ẩn danh',
            'SĐT người quyên góp': donation.donor_phone or '',
            'Email người quyên góp': donation.donor_email or '',
            'Số lượng': donation.quantity,
            'Đơn vị': donation.unit or 'cái',
            'Tình trạng': donation.condition or '',
            'Giá trị ước tính (VNĐ)': f"{donation.estimated_value:,.0f}" if donation.estimated_value else '',
            'Trạng thái': {
                'received': 'Đã nhận',
                'distributed': 'Đã phân phối',
                'damaged': 'Hư hỏng',
                'lost': 'Mất'
            }.get(donation.status, donation.status),
            'Ngày quyên góp': donation.donation_date.strftime('%d/%m/%Y'),
            'Người nhận phân phối': donation.recipient_name or '',
            'Loại phân phối': {
                'individual': 'Cá nhân',
                'class': 'Lớp học',
                'other': 'Khác'
            }.get(donation.recipient_type, donation.recipient_type) if donation.recipient_type else '',
            'Ngày phân phối': donation.distribution_date.strftime('%d/%m/%Y') if donation.distribution_date else '',
            'Ghi chú phân phối': donation.distribution_notes or '',
            'Người tạo': donation.creator.full_name if donation.creator else '',
            'Ngày tạo': donation.created_at.strftime('%d/%m/%Y %H:%M') if donation.created_at else ''
        })

    filename = f'tai_san_quyen_gop_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Tài sản quyên góp')

def export_events_to_excel(events):
    """
    Export events data to Excel
    """
    data = []
    for event in events:
        data.append({
            'ID': event.id,
            'Tiêu đề': event.title,
            'Mô tả': event.description or '',
            'Ngày bắt đầu': event.start_date.strftime('%d/%m/%Y %H:%M'),
            'Ngày kết thúc': event.end_date.strftime('%d/%m/%Y %H:%M') if event.end_date else '',
            'Địa điểm': event.location or '',
            'Loại sự kiện': event.event_type or '',
            'Trạng thái': {
                'planned': 'Đã lên kế hoạch',
                'ongoing': 'Đang diễn ra',
                'completed': 'Đã hoàn thành',
                'cancelled': 'Đã hủy'
            }.get(event.status, event.status),
            'Người tạo': event.creator.full_name if event.creator else '',
            'Ngày tạo': event.created_at.strftime('%d/%m/%Y %H:%M') if event.created_at else ''
        })

    filename = f'su_kien_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(data, filename, 'Sự kiện')
